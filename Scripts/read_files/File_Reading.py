from Scripts.read_files.Data_Parser import file_name_match, Read_files
from Scripts.read_files.Convert_Data import data_conversion
from Scripts.read_files.Generate_Spreadsheet import generate_spreadsheet
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


def reading_files(pstat, Area, loading, check_load, checkbutton, Shift, Convert, Ref, Notes):
    from Scripts.read_files import Files_Selection
    global spreadsheet
#    status = Label(root, text='')
#    status.place(x=stat_x, y =stat_y)
#    status.destroy()
#    status = Label(root, text='Status: Running')
#    status.place(x=stat_x, y =stat_y)
    files = Files_Selection.filenames
    directory = files[0]
    directory_2 = os.path.split(directory)[0]
    spreadsheet = Workbook()
    potentiostat = str(pstat.get())
    if potentiostat == 'Nova':
        count = 0
        for i in files:
            name = file_name_match(i)
            name = name+'_'+str(count)
            data = Read_files(i)
            data = data.read_Nova_files()
            ref_check = checkbutton.get()
            if 'Potential' in data.columns:
                if ref_check == 1:
                    shift = Shift.get()
                    convert = Convert.get()
                    ref = Ref.get()
                    data = data_conversion(data)
                    data = data.ref_conversion(shift, convert, ref)
                else:
                    new = 'Potential / V'
                    data.columns = [new if x=='Potential' else x for x in data.columns]
            
            if 'Current' in data.columns:
                res = Area.get()
                if len(res) != 0:
                    data = data_conversion(data)
                    data = data.current_density_convert(res)
                    load_check = check_load.get()
                    if load_check == 1:
                        load_entry = loading.get()
                        if len(load_entry) != 0:
                            data = data_conversion(data)
                            data = data.loading_convert(load_entry, res)
                else:
                    new = 'Current / A'
                    data.columns = [new if x=='Current' else x for x in data.columns]
            if len(Notes.get('1.0', 'end-1c')) != 0:
                nots = Notes.get('1.0', 'end-1c')
                if count == 0:
                    ws = generate_spreadsheet(spreadsheet, data)
                    ws = ws.Notes_page(nots)
                    count=count+1

            ws = spreadsheet.active
            if count == 0:
                ws.title = name
            if count > 0:
                ws = spreadsheet.create_sheet(name)
            ws = spreadsheet[str(name)]

            for r in dataframe_to_rows(data, index=True, header=True):
                ws.append(r)
            for cell in ws['A'] + ws[1]:
                cell.style = 'Pandas'
            count = count + 1
#        status.destroy()
#        status = Label(root, text='Status: Spreadsheet Made')
#        status.place(x=stat_x, y =stat_y)
    if potentiostat == 'EC-Lab':
        count=0
        for i in files:
            name = file_name_match(i)
            name = name+'_'+str(count)
            data = Read_files(i)
            data = data.read_EC_lab_files()
            ref_check = checkbutton.get()           
            if 'Ewe/V' in data.columns:
                if ref_check == 1:
                    shift = Shift.get()
                    convert = Convert.get()
                    ref = Ref.get()
                    data = data_conversion(data)
                    data = data.ref_conversion(shift, convert, ref)
                else:
                    new = 'Potential / V'
                    data.columns = [new if x=='Ewe/V' else x for x in data.columns]
            if '<I>/mA' in data.columns:
                res = Area.get()
                if len(res) != 0:
                    data = data_conversion(data)
                    data = data.current_density_convert(res)
                    load_check = check_load.get()
                    if load_check == 1:
                        load_entry = loading.get()
                        if len(load_entry) != 0:
                            data = data_conversion(data)
                            data = data.loading_convert(load_entry, res)
                else:
                    new = 'Current / mA'
                    data.columns = [new if x=='<I>/mA' else x for x in data.columns]
            
            if len(Notes.get('1.0', 'end-1c')) != 0:
                nots = Notes.get('1.0', 'end-1c')
                if count == 0:
                    ws = generate_spreadsheet(spreadsheet, data)
                    ws = ws.Notes_page(nots)
                    count=count+1
           
            ws = spreadsheet.active
            if count == 0:
                ws.title = name
            if count > 0:
                ws = spreadsheet.create_sheet(name)
            ws = spreadsheet[str(name)]

            for r in dataframe_to_rows(data, index=True, header=True):
                ws.append(r)
            for cell in ws['A'] + ws[1]:
                cell.style = 'Pandas'
            count = count + 1
#        status.destroy()
#        status = Label(root, text='Status: Spreadsheet Made')
#        status.place(x=stat_x, y =stat_y)
    if potentiostat == 'Ivium':
        pass
 #       status.destroy()
 #       status = Label(root, text='Status: Ivium parcer not made yet')
 #       status.place(x=stat_x, y =stat_y)