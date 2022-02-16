from tkinter import *
from tkinter import ttk
import tkinter as tk
from Scripts.analysis import plot_data
from openpyxl import load_workbook
from Scripts.Analysis_Window_Check import *
from Scripts.Plot_Window import *
from Scripts.Integrate_Window import *
thisdir = os.path.dirname(__file__)
rcfile = os.path.join(thisdir, '..', 'Misc', 'Icon.ico')

class analysis_window:
    def __init__(self, root):
        from Scripts.read_files import Loading_Excel, Save_Files
        self.root = root

        analysis_Window = Toplevel(self.root)
        analysis_Window.title("Analysis")
        analysis_Window.geometry("800x260")
        analysis_Window.resizable(0, 0)
        analysis_Window.iconbitmap(bitmap=rcfile)

        sheet_select = Label(analysis_Window, text="Sheet Selection")
        sheet_select.place(x=10, y=10)
        plotting_label = Label(analysis_Window, text="Plotting Parameters")
        plotting_label.place(x=250, y=10)
        style = ttk.Style()
        style.configure("BW.TLabel", background="white")
        container = ttk.Frame(analysis_Window, style="BW.TLabel")
        right_frame = ttk.Frame(analysis_Window, style="BW.TLabel", width=600, height=200)

        canvas = tk.Canvas(container, bg='white', width=200)
        right_frame_canvas = tk.Canvas(right_frame, bg='white', width=400)
        scrollbar_3 = ttk.Scrollbar(container, orient="horizontal", command=canvas.xview)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        right_scrollbar = ttk.Scrollbar(right_frame, orient='vertical', command=right_frame_canvas.yview)
        right_scrollbar_x = ttk.Scrollbar(right_frame, orient='horizontal', command=right_frame_canvas.xview)
        
        canvas.configure(xscrollcommand=scrollbar_3.set, width=200, heigh=200) 
        canvas.configure(yscrollcommand=scrollbar.set, width=200, heigh=200)
        right_frame_canvas.configure(yscrollcommand=right_scrollbar.set, width=400, heigh=200)
        right_frame_canvas.configure(xscrollcommand=right_scrollbar_x.set, width=400, heigh=200)

        scrollbar.pack(side="right", fill="y")
        right_scrollbar.pack(side="right", fill="y")
        right_scrollbar_x.pack(side="bottom", fill="x")
        scrollbar_3.pack(side="bottom", fill="x")

        right_scrollable_frame=ttk.Frame(right_frame_canvas, style="BW.TLabel")
        scrollable_frame = ttk.Frame(canvas, style="BW.TLabel")

        right_scrollable_frame.bind("<Configure>", lambda e: right_frame_canvas.configure(scrollregion=right_frame_canvas.bbox("all")))
        scrollable_frame.bind("<Configure>",lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        right_frame_canvas.create_window((0,0), window=right_scrollable_frame, anchor='nw')
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        Excel_label = Label(right_scrollable_frame, text="Sheet Name", bg="white", font='Helvetica 9 bold')
        Excel_label.grid(row=0, column=0)
        x_label = Label(right_scrollable_frame, text="x Axis", bg="white", font='Helvetica 9 bold')
        x_label.grid(row=0, column=1)
        y_label = Label(right_scrollable_frame, text="y Axis", bg="white", font='Helvetica 9 bold')
        y_label.grid(row=0, column=2)
        z_label = Label(right_scrollable_frame, text='z Axis', bg="white", font='Helvetica 9 bold')
        z_label.grid(row=0, column=3) 
        scan_label = Label(right_scrollable_frame, text='Scan', bg="white", font='Helvetica 9 bold')
        scan_label.grid(row=0, column=4)

        check_z = IntVar(value=0)
        def FlipState_2():
            global State
            State = check_z.get()
                
        Check_Z = ttk.Checkbutton(analysis_Window, text='Colour By Z', variable=check_z, onvalue='1', offvalue='0', command=FlipState_2)
        Check_Z.place(x=680, y=150)

        #scan_selection
        Scan_select = Label(analysis_Window, text='Scan:')
        Scan_select.place(x=670, y=90)
        scan_select = Entry(analysis_Window, width=8, borderwidth=5)
        scan_select.place(x=720, y=90)

        try:    
            save_loc = Loading_Excel.save_loc
        except:
            save_loc = Save_Files.save_loc 
        
        wb = load_workbook(filename=save_loc)
        sheets = wb.sheetnames
        files = []
        for i in sheets:
            files.append(i)

        Plot_button = ttk.Button(analysis_Window, text="Plot", command=lambda: plot_window(self.root, files, wb, check_z, scan_select))
        Plot_button.place(x=690, y=30)

        integrate_button = ttk.Button(analysis_Window, text="Integrate", command=lambda: integrate_window(self.root, check_z, files, wb, sheets))
        integrate_button.place(x=690, y=60)

        cbTexts={}
        cbVariables={}
        cb={}
        count = 0
        for i in files:
            cbTexts[i] = StringVar()
            cbTexts[i].set(i)
            cbVariables[count] = IntVar()
            cbVariables[count].set(0)
            cb[count] = Checkbutton(scrollable_frame, textvariable=cbTexts[i], variable=cbVariables[count], onvalue='1', offvalue='0',  command=lambda: check_state(cb, cbVariables, wb, files, right_scrollable_frame), bg='white')
            cb[count].pack(anchor="w")
            count=count+1
        container.place(x=10, y=30)
        right_frame.place(x=250, y=30)

        canvas.pack(side="left", fill="both", expand=True)
        right_frame_canvas.pack(side="right", fill="both", expand=True)